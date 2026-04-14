"""Скрипт экспорта предсказаний из БД Coffee Oracle в Excel.

Использование:
    python export_predictions.py
    python export_predictions.py --db /путь/к/coffee_oracle.db
    python export_predictions.py --db ./data/coffee_oracle.db --out ./exports/predictions.xlsx
"""

import argparse
import os
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─── Константы ───────────────────────────────────────────────────────────────

ЗАГОЛОВКИ = [
    "ID предсказания",
    "Дата и время",
    "Платформа",
    "Telegram ID",
    "Username",
    "Полное имя",
    "Email",
    "Тип подписки",
    "Запрос пользователя",
    "Текст предсказания",
]

ШИРИНЫ_КОЛОНОК = [16, 20, 12, 16, 20, 24, 28, 16, 40, 60]

ПЛАТФОРМЫ = {"tg": "Telegram", "max": "MAX"}

# Цвета шапки таблицы
ЦВЕТ_ШАПКИ = "2F4F8F"       # тёмно-синий фон
ЦВЕТ_ТЕКСТА_ШАПКИ = "FFFFFF" # белый текст

# Цвета чередующихся строк
ЦВЕТ_СТРОКИ_ЧЁТНОЙ = "EEF2FF"
ЦВЕТ_СТРОКИ_НЕЧЁТНОЙ = "FFFFFF"


# ─── Получение пути к БД ─────────────────────────────────────────────────────

def получить_путь_к_бд(аргумент_db: str | None) -> Path:
    """Определяет путь к файлу БД из аргумента командной строки или .env."""
    if аргумент_db:
        путь = Path(аргумент_db)
    else:
        load_dotenv()
        db_name = os.getenv("DB_NAME", "coffee_oracle.db")
        # Ищем БД сначала рядом со скриптом, потом в data/
        варианты = [
            Path(db_name),
            Path("data") / db_name,
            Path("app") / "data" / db_name,
        ]
        путь = None
        for вариант in варианты:
            if вариант.exists():
                путь = вариант
                break
        if путь is None:
            путь = Path(db_name)  # вернём как есть — ошибка будет ниже

    if not путь.exists():
        print(f"[ОШИБКА] Файл базы данных не найден: {путь}")
        print("Укажи путь явно: python export_predictions.py --db /путь/к/coffee_oracle.db")
        sys.exit(1)

    return путь


# ─── Запрос к БД ─────────────────────────────────────────────────────────────

ЗАПРОС = """
    SELECT
        p.id                  AS prediction_id,
        p.created_at          AS created_at,
        u.source              AS source,
        u.telegram_id         AS telegram_id,
        u.username            AS username,
        u.full_name           AS full_name,
        u.email               AS email,
        p.subscription_type   AS subscription_type,
        p.user_request        AS user_request,
        p.prediction_text     AS prediction_text
    FROM predictions p
    JOIN users u ON u.id = p.user_id
    WHERE u.deleted_at IS NULL
    ORDER BY p.created_at DESC
"""


def загрузить_данные(путь_к_бд: Path) -> list[tuple]:
    """Выполняет запрос к SQLite и возвращает список строк."""
    try:
        соединение = sqlite3.connect(путь_к_бд)
        курсор = соединение.cursor()
        курсор.execute(ЗАПРОС)
        строки = курсор.fetchall()
        соединение.close()
        return строки
    except sqlite3.OperationalError as ошибка:
        print(f"[ОШИБКА] Не удалось выполнить запрос к БД: {ошибка}")
        sys.exit(1)


# ─── Форматирование значений ──────────────────────────────────────────────────

def форматировать_дату(значение: str | None) -> str:
    """Преобразует ISO-строку даты в читаемый формат."""
    if not значение:
        return ""
    try:
        # SQLite хранит даты в формате "2024-01-15 12:34:56" или ISO 8601
        для_парсинга = значение.replace("T", " ").split(".")[0].split("+")[0]
        дата = datetime.strptime(для_парсинга, "%Y-%m-%d %H:%M:%S")
        return дата.strftime("%d.%m.%Y %H:%M")
    except (ValueError, AttributeError):
        return str(значение)


def форматировать_платформу(значение: str | None) -> str:
    """Преобразует код платформы в читаемое название."""
    return ПЛАТФОРМЫ.get(значение or "", значение or "")


def подготовить_строку(строка: tuple) -> list:
    """Преобразует кортеж из БД в список ячеек для Excel."""
    (
        prediction_id, created_at, source, telegram_id,
        username, full_name, email, subscription_type,
        user_request, prediction_text,
    ) = строка

    return [
        prediction_id,
        форматировать_дату(created_at),
        форматировать_платформу(source),
        telegram_id,
        username or "",
        full_name or "",
        email or "",
        subscription_type or "free",
        user_request or "",
        prediction_text or "",
    ]


# ─── Создание Excel ───────────────────────────────────────────────────────────

def создать_excel(строки: list[tuple], путь_к_файлу: Path) -> None:
    """Создаёт и сохраняет Excel-файл с данными предсказаний."""
    wb = Workbook()
    лист = wb.active
    лист.title = "Предсказания"

    # Заморозка первой строки (шапка всегда видна при прокрутке)
    лист.freeze_panes = "A2"

    # ── Шапка таблицы ────────────────────────────────────────────────────────
    шрифт_шапки = Font(
        name="Arial", bold=True, color=ЦВЕТ_ТЕКСТА_ШАПКИ, size=10
    )
    заливка_шапки = PatternFill(
        fill_type="solid", start_color=ЦВЕТ_ШАПКИ, end_color=ЦВЕТ_ШАПКИ
    )
    выравнивание_шапки = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    for номер, заголовок in enumerate(ЗАГОЛОВКИ, start=1):
        ячейка = лист.cell(row=1, column=номер, value=заголовок)
        ячейка.font = шрифт_шапки
        ячейка.fill = заливка_шапки
        ячейка.alignment = выравнивание_шапки

    лист.row_dimensions[1].height = 32

    # ── Данные ───────────────────────────────────────────────────────────────
    шрифт_данных = Font(name="Arial", size=9)
    выравнивание_текст = Alignment(vertical="top", wrap_text=True)
    выравнивание_центр = Alignment(horizontal="center", vertical="top")

    # Колонки, которые выравниваем по центру
    центрированные_колонки = {1, 2, 3, 4, 8}

    for номер_строки, строка in enumerate(строки, start=2):
        данные = подготовить_строку(строка)

        # Чередование цвета строк
        цвет = ЦВЕТ_СТРОКИ_ЧЁТНОЙ if номер_строки % 2 == 0 else ЦВЕТ_СТРОКИ_НЕЧЁТНОЙ
        заливка = PatternFill(fill_type="solid", start_color=цвет, end_color=цвет)

        for номер_колонки, значение in enumerate(данные, start=1):
            ячейка = лист.cell(row=номер_строки, column=номер_колонки, value=значение)
            ячейка.font = шрифт_данных
            ячейка.fill = заливка
            if номер_колонки in центрированные_колонки:
                ячейка.alignment = выравнивание_центр
            else:
                ячейка.alignment = выравнивание_текст

    # ── Ширина колонок ────────────────────────────────────────────────────────
    for номер, ширина in enumerate(ШИРИНЫ_КОЛОНОК, start=1):
        лист.column_dimensions[get_column_letter(номер)].width = ширина

    # ── Автофильтр ────────────────────────────────────────────────────────────
    лист.auto_filter.ref = (
        f"A1:{get_column_letter(len(ЗАГОЛОВКИ))}1"
    )

    # ── Итоговый лист со статистикой ─────────────────────────────────────────
    лист_статистики = wb.create_sheet("Статистика")
    _заполнить_лист_статистики(лист_статистики, строки)

    # ── Сохранение ────────────────────────────────────────────────────────────
    путь_к_файлу.parent.mkdir(parents=True, exist_ok=True)
    wb.save(путь_к_файлу)


def _заполнить_лист_статистики(лист, строки: list[tuple]) -> None:
    """Заполняет лист со сводной статистикой."""
    шрифт_заголовок = Font(name="Arial", bold=True, size=11)
    шрифт_обычный = Font(name="Arial", size=10)
    заливка_заголовок = PatternFill(fill_type="solid", start_color=ЦВЕТ_ШАПКИ, end_color=ЦВЕТ_ШАПКИ)
    шрифт_белый_жирный = Font(name="Arial", bold=True, color="FFFFFF", size=11)

    def шапка(строка, кол, текст):
        яч = лист.cell(row=строка, column=кол, value=текст)
        яч.font = шрифт_белый_жирный
        яч.fill = заливка_заголовок
        яч.alignment = Alignment(horizontal="center", vertical="center")

    def значение(строка, кол, текст):
        яч = лист.cell(row=строка, column=кол, value=текст)
        яч.font = шрифт_обычный
        яч.alignment = Alignment(horizontal="center", vertical="center")

    # Подсчёт метрик
    всего = len(строки)
    по_платформе: dict[str, int] = {}
    по_подписке: dict[str, int] = {}
    уникальные_пользователи: set = set()

    for строка in строки:
        _, _, source, telegram_id, _, _, _, subscription_type, _, _ = строка
        платформа = форматировать_платформу(source)
        по_платформе[платформа] = по_платформе.get(платформа, 0) + 1
        подписка = subscription_type or "free"
        по_подписке[подписка] = по_подписке.get(подписка, 0) + 1
        уникальные_пользователи.add((telegram_id, source))

    # ── Вывод ────────────────────────────────────────────────────────────────
    яч = лист.cell(row=1, column=1, value="Сводная статистика предсказаний")
    яч.font = Font(name="Arial", bold=True, size=13)
    яч.alignment = Alignment(horizontal="center")
    лист.merge_cells("A1:B1")

    шапка(3, 1, "Показатель")
    шапка(3, 2, "Значение")
    значение(4, 1, "Всего предсказаний")
    значение(4, 2, всего)
    значение(5, 1, "Уникальных пользователей")
    значение(5, 2, len(уникальные_пользователи))

    строка_вывода = 7
    яч = лист.cell(row=строка_вывода, column=1, value="По платформам")
    яч.font = шрифт_заголовок
    строка_вывода += 1
    шапка(строка_вывода, 1, "Платформа")
    шапка(строка_вывода, 2, "Количество")
    строка_вывода += 1
    for платформа, кол in sorted(по_платформе.items()):
        значение(строка_вывода, 1, платформа)
        значение(строка_вывода, 2, кол)
        строка_вывода += 1

    строка_вывода += 1
    яч = лист.cell(row=строка_вывода, column=1, value="По типу подписки")
    яч.font = шрифт_заголовок
    строка_вывода += 1
    шапка(строка_вывода, 1, "Подписка")
    шапка(строка_вывода, 2, "Количество")
    строка_вывода += 1
    for подписка, кол in sorted(по_подписке.items()):
        значение(строка_вывода, 1, подписка)
        значение(строка_вывода, 2, кол)
        строка_вывода += 1

    лист.column_dimensions["A"].width = 30
    лист.column_dimensions["B"].width = 16


# ─── Точка входа ─────────────────────────────────────────────────────────────

def main() -> None:
    """Точка входа скрипта."""
    парсер = argparse.ArgumentParser(
        description="Экспорт предсказаний Coffee Oracle в Excel"
    )
    парсер.add_argument(
        "--db",
        type=str,
        default=None,
        help="Путь к файлу SQLite БД (по умолчанию берётся из .env → DB_NAME)",
    )
    парсер.add_argument(
        "--out",
        type=str,
        default=None,
        help="Путь к выходному Excel-файлу (по умолчанию: exports/predictions_ДАТА.xlsx)",
    )
    аргументы = парсер.parse_args()

    # Путь к БД
    путь_к_бд = получить_путь_к_бд(аргументы.db)

    # Путь к выходному файлу
    если_не_указан = f"exports/predictions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    путь_к_excel = Path(аргументы.out) if аргументы.out else Path(если_не_указан)

    print(f"[INFO] Подключаемся к БД: {путь_к_бд}")
    строки = загрузить_данные(путь_к_бд)

    if not строки:
        print("[INFO] В базе данных нет предсказаний. Файл не создан.")
        sys.exit(0)

    print(f"[INFO] Найдено предсказаний: {len(строки)}")
    print(f"[INFO] Создаём Excel-файл: {путь_к_excel}")

    создать_excel(строки, путь_к_excel)

    print(f"[OK] Готово! Файл сохранён: {путь_к_excel}")
    print(f"[OK] Строк в таблице: {len(строки)}")


if __name__ == "__main__":
    main()
